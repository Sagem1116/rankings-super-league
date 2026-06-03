from pathlib import Path
import sys
try:
    import openpyxl
except Exception as e:
    print('NO_OPENPYXL', e)
    sys.exit(0)
path = Path('teste.xlsx')
if not path.exists():
    print('NO_FILE')
    sys.exit(0)
wb = openpyxl.load_workbook(path, read_only=True)
for name in wb.sheetnames:
    ws = wb[name]
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    print(name, row)
